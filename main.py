# This is a sample Python script.
import json

import pymysql
# Press Shift+F10 to execute it or replace it with your code.
# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.
import requests
from bs4 import BeautifulSoup
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Alignment
from requests.adapters import HTTPAdapter
import re
import openpyxl
from openpyxl.styles.numbers import NumberFormat


def print_hi(name):
    # Use a breakpoint in the code line below to debug your script.
    # 发送请求获取HTML页面
    url = "https://hq.smm.cn/copper/fullscreen"
    response = requests.get(url)
    html = response.text

    # 使用Beautiful Soup解析HTML
    soup = BeautifulSoup(html, "html.parser")

    # 提取a标签的href属性和alt属性的值
    data = []
    for a_tag in soup.find_all("a", class_="category-li-a"):
        href = a_tag.get("href")
        alt = a_tag.get("alt")
        data.append({"大类类别": alt, "url": href})

    # 将数据转换为DataFrame
    df = pd.DataFrame(data)
    df.to_csv('data.csv', index=False)
    # 打印DataFrame
    print(df)


# Press the green button in the gutter to run the script.
def print_csv():
    # 发送请求获取网页内容
    url = 'https://hq.smm.cn/copper/fullscreen'
    response = requests.get(url)
    html_content = response.text

    # 使用BeautifulSoup解析HTML
    soup = BeautifulSoup(html_content, 'html.parser')

    # 提取a标签的href和alt属性值
    data = []
    for a_tag in soup.find_all('a', class_='category-li-a'):
        href = a_tag['href']
        alt = a_tag['alt']

        # 拼接完整URL
        full_url = href + '/fullscreen'
        data.append([alt, href, full_url])
    #  data.append([alt,href])

    # 将数据存储到CSV文件中
    df = pd.DataFrame(data, columns=['类别', 'url', '类别详情url'])
    df.to_csv('D:\工作记录\dcost\成本网站特点分析\data1.csv', index=False)

    print('数据已存储到 data.csv 文件中。')


def print_csv1():
    # 发送请求获取网页内容
    url = 'https://hq.smm.cn/copper/fullscreen'
    response = requests.get(url)
    html_content = response.text

    # 使用BeautifulSoup解析HTML
    soup = BeautifulSoup(html_content, 'html.parser')

    # 提取a标签的href和alt属性值
    data = []
    categories = set()  # 用于存储类别详情的set集合

    for a_tag in soup.find_all('a', class_='category-li-a'):
        href = a_tag['href']
        alt = a_tag['alt']
        # 拼接完整URL
        full_url = href + '/fullscreen'
        data.append([full_url, alt])
        categories.add(full_url)  # 将类别详情添加到set集合

    # 循环访问类别详情的URL，并从中提取详情项的URL和名称
    for category in categories:
        response = requests.get(category)
        html_content = response.text
        soup = BeautifulSoup(html_content, 'html.parser')
        for td_tag in soup.find_all('td', class_='product-name'):
            a_tag = td_tag.find('a')
            if a_tag is not None:  # 只处理含有子节点<a>标签的<td>标签
                href = 'https://hq.smm.cn' + a_tag['href']
                text = a_tag.text
                data.append([href, text])

    # 将数据存储到CSV文件中
    df = pd.DataFrame(data, columns=['详情项', '项名称'])
    df.to_csv('D:\工作记录\dcost\成本网站特点分析\data2.csv', index=False)

    print('数据已存储到 data.csv 文件中。')


def print_csv3():
    # 发送请求获取网页内容
    url = 'https://hq.smm.cn/copper/fullscreen'
    response = requests.get(url)
    html_content = response.text

    # 使用BeautifulSoup解析HTML
    soup = BeautifulSoup(html_content, 'html.parser')

    # 提取a标签的href和alt属性值
    category_data = []  # 用于存储类别详情的列表，每个元素为一个字典，包含大类名称、大类URL、类别名称和类别URL

    for a_tag in soup.find_all('a', class_='category-li-a'):
        category_name = a_tag['alt']
        category_url = a_tag['href'] + '/fullscreen'
        category_data.append({'大类名称': category_name, '大类URL': category_url})

    # 循环访问类别详情的URL，并从中提取详情项的URL和名称，并按类别进行分类
    data = []
    for category in category_data:
        category_name = category['大类名称']
        category_url = category['大类URL']
        response = requests.get(category_url)
        html_content = response.text
        soup = BeautifulSoup(html_content, 'html.parser')
        for td_tag in soup.find_all('td', class_='product-name'):
            a_tag = td_tag.find('a')
            if a_tag is not None:  # 只处理含有子节点<a>标签的<td>标签
                href = 'https://hq.smm.cn' + a_tag['href']
                text = a_tag.text
                data.append([category_name, category_url, text, href])

    # 将数据存储到CSV文件中
    df = pd.DataFrame(data, columns=['大类名称', '大类URL', '项名称', '详情项'])
    df.to_csv('D:\工作记录\dcost\成本网站特点分析\data3.csv', index=False)

    print('数据已存储到 data3.csv 文件中。')


def print_csv4():
    # 发送请求获取网页内容
    url = 'https://hq.smm.cn/copper/fullscreen'
    response = requests.get(url)
    html_content = response.text
    # 计算时间范围
    end_date = datetime.now().strftime('%Y-%m-%d')  # 当前时间
    start_date = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')  # 当前时间前30天
    # 使用BeautifulSoup解析HTML
    soup = BeautifulSoup(html_content, 'html.parser')

    # 提取a标签的href和alt属性值
    category_data = []  # 用于存储类别详情的列表，每个元素为一个字典，包含大类名称、大类URL、类别名称和类别URL

    for a_tag in soup.find_all('a', class_='category-li-a'):
        category_name = a_tag['alt']
        category_url = a_tag['href'] + '/fullscreen'
        category_data.append({'大类名称': category_name, '大类URL': category_url})

    # 循环访问类别详情的URL，并从中提取详情项的URL和名称，并按类别进行分类
    data = []
    for category in category_data:
        category_name = category['大类名称']
        category_url = category['大类URL']
        response = requests.get(category_url)
        html_content = response.text
        soup = BeautifulSoup(html_content, 'html.parser')
        for td_tag in soup.find_all('td', class_='product-name'):
            a_tag = td_tag.find('a')
            if a_tag is not None:  # 只处理含有子节点<a>标签的<td>标签
                href = 'https://hq.smm.cn' + a_tag['href']
                category_id = a_tag['href'].split('/category/')[1]
                text = a_tag.text.replace(' ', '').replace('\t', '').replace('\n', '')  # 替换所有空白字符为空字符串
                url_detail = 'https://hq.smm.cn/ajax/spot/history/' + category_id + '/' + start_date + '/' + end_date
                data.append({'大类名称': category_name, '大类URL': category_url, '项名称': text, '详情项': href,
                             '分类ID': category_id + '', "30天详细url": url_detail})

    # 将数据存储到CSV文件中
    df = pd.DataFrame(data)
    df.to_csv(r'D:\工作记录\dcost\成本网站特点分析\data6.csv', index=False, float_format='%.0f')

    print('数据已存储到 data.csv 和 data.json 文件中。')


def print_save_sql():
    # 发送请求获取网页内容
    url = 'https://hq.smm.cn/copper/fullscreen'
    response = requests.get(url)
    html_content = response.text
    # 计算时间范围
    end_date = datetime.now().strftime('%Y-%m-%d')  # 当前时间
    start_date = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')  # 当前时间前30天

    # 使用BeautifulSoup解析HTML
    soup = BeautifulSoup(html_content, 'html.parser')

    # 提取a标签的href和alt属性值
    category_data = []  # 用于存储类别详情的列表，每个元素为一个字典，包含大类名称、大类URL、类别名称和类别URL

    for a_tag in soup.find_all('a', class_='category-li-a'):
        category_name = a_tag['alt']
        category_url = a_tag['href'] + '/fullscreen'
        category_data.append({'大类名称': category_name, '大类URL': category_url})

    # 循环访问类别详情的URL，并从中提取详情项的URL和名称，并按类别进行分类
    data = []
    for category in category_data:
        category_name = category['大类名称']
        category_url = category['大类URL']
        response = requests.get(category_url)
        html_content = response.text
        soup = BeautifulSoup(html_content, 'html.parser')
        for td_tag in soup.find_all('td', class_='product-name'):
            a_tag = td_tag.find('a')
            if a_tag is not None:  # 只处理含有子节点<a>标签的<td>标签
                href = 'https://hq.smm.cn/ajax/spot/history/' + a_tag['href'].split('/category/')[1]
                text = a_tag.text.replace(' ', '').replace('\t', '').replace('\n', '')  # 替换所有空白字符为空字符串
                category_id = a_tag['href'].split('/category/')[1]
                url_detail = 'https://hq.smm.cn/ajax/spot/history/' + category_id + '/' + start_date + '/' + end_date
                data.append((category_name, category_url, text, href, category_id, url_detail))

    # 连接MySQL数据库
    conn = pymysql.connect(host='localhost', user='root', password='123456', db='docost-data',
                           charset='utf8mb4')
    cursor = conn.cursor()

    # 创建数据表
    create_table_sql = '''
    CREATE TABLE IF NOT EXISTS smm_data (
      id INT AUTO_INCREMENT PRIMARY KEY,
      大类名称 VARCHAR(255),
      大类URL VARCHAR(255),
      项名称 VARCHAR(255),
      详情项 VARCHAR(255),
      分类ID VARCHAR(255),
      30天详细项目分类IdUrl VARCHAR(255)
    )
    '''
    cursor.execute(create_table_sql)

    # 插入数据到数据表
    insert_data_sql = '''
    INSERT INTO smm_data (大类名称, 大类URL, 项名称, 详情项, 分类ID,30天详细项目分类IdUrl)
    VALUES (%s, %s, %s, %s, %s,%s)
    '''
    cursor.executemany(insert_data_sql, data)
    conn.commit()

    # 关闭数据库连接
    cursor.close()
    conn.close()

    print('数据已成功存储到MySQL数据库中。')


def print_excel_you_se():
    # 发送请求获取网页内容
    url = 'https://hq.smm.cn/copper/fullscreen'
    response = requests.get(url)
    html_content = response.text

    # 使用BeautifulSoup解析HTML
    soup = BeautifulSoup(html_content, 'html.parser')

    # 提取a标签的href和alt属性值
    category_data = []  # 用于存储类别详情的列表，每个元素为一个字典，包含大类名称、大类URL、类别名称和类别URL
    # 计算时间范围
    end_date = datetime.now().strftime('%Y-%m-%d')  # 当前时间
    start_date = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')  # 当前时间前30天
    for a_tag in soup.find_all('a', class_='category-li-a'):
        category_name = a_tag['alt']
        category_url = a_tag['href'] + '/fullscreen'
        category_data.append({'大类名称': category_name, '大类URL': category_url})

    # 循环访问类别详情的URL，并从中提取详情项的URL和名称，并按类别进行分类
    data = []
    for category in category_data:
        category_name = category['大类名称']
        category_url = category['大类URL']
        response = requests.get(category_url)
        html_content = response.text
        soup = BeautifulSoup(html_content, 'html.parser')
        for td_tag in soup.find_all('td', class_='product-name'):
            a_tag = td_tag.find('a')
            if a_tag is not None:  # 只处理含有子节点<a>标签的<td>标签
                href = 'https://hq.smm.cn' + a_tag['href']
                category_id = a_tag['href'].split('/category/')[1]
                text = a_tag.text.replace(' ', '').replace('\t', '').replace('\n', '')  # 替换所有空白字符为空字符串
                url_detail = 'https://hq.smm.cn/ajax/spot/history/' + category_id + '/' + start_date + '/' + end_date
                data.append({'大类名称': category_name, '大类URL': category_url, '项目名称': text, '项目url': href,
                             '项目详细分类ID': category_id + '', "30天详细项目分类IdUrl": url_detail})

    # 将数据存储到Excel文件中
    df = pd.DataFrame(data)
    df.to_excel('D:\\工作记录\\dcost\\成本网站特点分析\\data6.xlsx', index=False)

    print('数据已存储到 data.xlsx 文件中。')
    # 将数据存储到JSON文件中，指定编码方式为utf-8
    with open(r'D:\工作记录\dcost\成本网站特点分析\data.json', 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False)

    print('数据已存储到 data.json 文件中。')


def print_excel_xin_neng_yuan():
    # 发送请求获取网页内容
    url = 'https://new-energy.smm.cn/new_energy/14042'
    page_url = 'https://new-energy.smm.cn'
    response = requests.get(url)
    html_content = response.text

    # 使用BeautifulSoup解析HTML
    soup = BeautifulSoup(html_content, 'html.parser')

    # 提取a标签的href和alt属性值
    category_data = []  # 用于存储类别详情的列表，每个元素为一个字典，包含大类名称、大类URL、类别名称和类别URL
    # 计算时间范围
    end_date = datetime.now().strftime('%Y-%m-%d')  # 当前时间
    start_date = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')  # 当前时间前30天
    # 使用正则表达式匹配类名
    class_regex = re.compile(r'FilterTable_productName__\w+')  # 使用正则表达式匹配类名
    for a_tag in soup.find_all('a', class_='Nav_groupColumn__CJqRh'):
        if a_tag is not None:  # 只处理含有子节点<a>标签的<td>标签
            category_name = a_tag.text
            category_url = page_url + a_tag['href']
            category_data.append({'大类名称': category_name, '大类URL': category_url})

    # 循环访问类别详情的URL，并从中提取详情项的URL和名称，并按类别进行分类
    data = []
    for category in category_data:
        category_name = category['大类名称']
        category_url = category['大类URL']
        response = requests.get(category_url)
        html_content = response.text
        soup = BeautifulSoup(html_content, 'html.parser')
        for a_tag in soup.find_all('a', class_=class_regex):
            if a_tag is not None:  # 只处理含有子节点<a>标签的<td>标签
                href = a_tag['href']
                category_id = a_tag['href'].split('/category/')[1]
                text = a_tag.text.replace(' ', '').replace('\t', '').replace('\n', '')  # 替换所有空白字符为空字符串
                url_detail = 'https://hq.smm.cn/ajax/spot/history/' + category_id + '/' + start_date + '/' + end_date
                data.append({'大类名称': category_name, '大类URL': category_url, '项目名称': text, '项目url': href,
                             '项目详细分类ID': category_id + '', "30天详细项目分类IdUrl": url_detail})

    # 将数据存储到Excel文件中
    df = pd.DataFrame(data)
    df.to_excel('D:\\工作记录\\dcost\\成本网站特点分析\\xin.xlsx', index=False)

    print('数据已存储到 xin.xlsx 文件中。')
    # 将数据存储到JSON文件中，指定编码方式为utf-8
    with open(r'D:\工作记录\dcost\成本网站特点分析\xin.json', 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False)

    print('数据已存储到 xin.json 文件中。')


def print_xin_nengyaun_guige():
    # 发送请求获取网页内容
    url = 'https://new-energy.smm.cn/new_energy/14042'
    page_url = 'https://new-energy.smm.cn'
    response = requests.get(url)
    html_content = response.text

    # 使用BeautifulSoup解析HTML
    soup = BeautifulSoup(html_content, 'html.parser')

    # 提取a标签的href和alt属性值
    category_data = []  # 用于存储类别详情的列表，每个元素为一个字典，包含大类名称、大类URL、类别名称和类别URL
    # 计算时间范围
    end_date = datetime.now().strftime('%Y-%m-%d')  # 当前时间
    start_date = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')  # 当前时间前30天
    # 使用正则表达式匹配类名
    class_regex = re.compile(r'FilterTable_productName__\w+')  # 使用正则表达式匹配类名
    for a_tag in soup.find_all('a', class_='Nav_groupColumn__CJqRh'):
        if a_tag is not None:  # 只处理含有子节点<a>标签的<td>标签
            category_name = a_tag.text
            category_url = page_url + a_tag['href']
            category_data.append({'大类名称': category_name, '大类URL': category_url})

    # 循环访问类别详情的URL，并从中提取详情项的URL和名称，并按类别进行分类
    data = []
    for category in category_data:
        category_name = category['大类名称']
        category_url = category['大类URL']
        response = requests.get(category_url)
        html_content = response.text
        soup = BeautifulSoup(html_content, 'html.parser')
        tbodys = soup.find_all('tbody', class_='ant-table-tbody')
        for tbody in tbodys:
            # Find all rows in the table
            rows = tbody.find_all('tr')
            # Loop through the rows and extract data
            for row in rows:
                a_tags = row.find_all('a', class_=class_regex)
                for a_tag in a_tags:
                    href = a_tag['href']
                    category_id = href.split('/category/')[1]
                    text = a_tag.text.replace(' ', '').replace('\t', '').replace('\n', '')  # 替换所有空白字符为空字符串
                    url_detail = 'https://hq.smm.cn/ajax/spot/history/' + category_id + '/' + start_date + '/' + end_date
                    td_tage = row.find_all('td')[1]
                    thickness_div = td_tage.find('div')
                    thickness = thickness_div.text.strip() if thickness_div else ''  # 判断是否存在 thickness_div，若存在则提取 text 属性并去除首尾空白字符，否则赋值为空字符串
                    data.append(
                        {'大类名称': category_name, '大类URL': category_url, '项目名称': text, '规格': thickness,
                         '项目url': href,
                         '项目详细分类ID': category_id + '', "30天详细项目分类IdUrl": url_detail})

    # 将数据存储到Excel文件中
    df = pd.DataFrame(data)
    df.to_excel(r'D:\工作记录\dcost\成本网站特点分析\in2.xlsx', index=False)

    print('数据已存储到 xin.xlsx 文件中。')
    # 将数据存储到JSON文件中，指定编码方式为utf-8
    with open(r'D:\工作记录\dcost\成本网站特点分析\xin2.json', 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False)

    print('数据已存储到 xin.json 文件中。')


def print_su_liao():
    headers1 = {
        'Accept': '*/*',
        'Accept-Encoding': 'gzip, deflate, br',
        'Accept-Language': 'zh-CN,zh;q=0.9',
        'Content-Type': 'application/json',
        'Cookie': 'guid=236fbd2c-7794-b7d1-6943-950e0b0b033b; __bid_n=1877924aad211066474207; accessId=30dbced0-f5cb-11eb-893a-df95eeb4af27; Hm_lvt_44c27e8e603ca3b625b6b1e9c35d712d=1681364658,1681723276,1681870152; BAIDU_SSP_lcr=https://www.bing.com/; route=258ceb4bb660681c2cb2768af9756936; ASP.NET_SessionId=0snh542onvy54tecbm2g3z1b; Hm_lvt_78a951b1e2ee23efdc6af2ce70d6b9be=1681723475,1681796427,1681870273; href=https%3A%2F%2Fprices.sci99.com%2Fcn%2Fproduct.aspx%3Fppid%3D12278%26ppname%3DLDPE%26navid%3D521; sensorsdata2015jssdkcross=%7B%22distinct_id%22%3A%2218779249e7f58c-0ef11c4dd18171-26031851-2073600-18779249e80dbd%22%2C%22first_id%22%3A%22%22%2C%22props%22%3A%7B%22%24latest_traffic_source_type%22%3A%22%E8%87%AA%E7%84%B6%E6%90%9C%E7%B4%A2%E6%B5%81%E9%87%8F%22%2C%22%24latest_search_keyword%22%3A%22%E6%9C%AA%E5%8F%96%E5%88%B0%E5%80%BC%22%2C%22%24latest_referrer%22%3A%22https%3A%2F%2Fwww.bing.com%2F%22%7D%2C%22%24device_id%22%3A%2218779249e7f58c-0ef11c4dd18171-26031851-2073600-18779249e80dbd%22%7D; Hm_lpvt_44c27e8e603ca3b625b6b1e9c35d712d=1681902304; FPTOKEN=icFK4xODoKnx/04rfAoITwuUJBg68likg2sSI4JgBPdIv/S9jhgix39inzUaoGz5hAIiczPUUZQ2uJszR6v//D7+LcTqx40Z/CUIznbrR2yIMrHsmyatlhH0WgIhPLKznwqKkRvr+eQeiSkLZK189FdYUoll7FikJ6lja/uPw8ovlkQniYTMSypuNrT6+UTfSY3c/zNlDEAUfVlT3B6V+SjmLK99mEfkUVbWRRy1qhZiXReWpiXusi+ouWZe7ga+D3dvXgY0ykCQpbimCHhFwdtqP2VuF4NlgxYEuCrqlInCraqusNR1YVevG3KVdQyAOqCGSXMRebnNUhDRf80qOeIpN4W9diPx64vABA6OmttX0NMtsw4+0GqgYg77xHsn0eVDVaNA7bNo117fWDYJDA==|SCjqHrqxrSGlMmf5y8HpUsQThVuV6tMmVwf2Q0PT0AE=|10|2701e0700fa2e0fe5c4736e6b0d909ed; STATReferrerIndexId=2; Hm_lpvt_78a951b1e2ee23efdc6af2ce70d6b9be=1681952969; qimo_seosource_30dbced0-f5cb-11eb-893a-df95eeb4af27=%E5%85%B6%E4%BB%96%E7%BD%91%E7%AB%99; qimo_seokeywords_30dbced0-f5cb-11eb-893a-df95eeb4af27=%E6%9C%AA%E7%9F%A5; pageViewNum=300',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/111.0.0.0 Safari/537.36'
    }
    res1 = requests.get('https://prices.sci99.com/api/nav/1', verify=False, timeout=10,
                        headers=headers1)
    # 将 JSON 数据解析成 Python 对象
    nav = json.loads(res1.text)
    result = []
    for data_items in nav['data']:
        if data_items['Name'] == '塑料':
            # 遍历数据，生成分类列表
            categories = []

            name1 = data_items['Name']
            for subcategory in data_items['Children']:
                name2 = subcategory['Name']
                for subsubcategory in subcategory['Children']:
                    name3 = subsubcategory['Name']
                    for subsubsubcategory in subsubcategory.get('Children', []):
                        name4 = subsubsubcategory['Name']
                        ppid = subsubsubcategory['Ppid']
                        navId = subsubsubcategory['ID']
                        result.append({
                            'name1': name1,
                            'name2': name2,
                            'name3': name3,
                            'name4': name4,
                            'Ppid': ppid,
                            'ID': navId,
                        })
    print(result)

    s = requests.Session()
    s.mount('http://', HTTPAdapter(max_retries=3))  # 访问http协议时，设置重传请求最多三次
    s.mount('https://', HTTPAdapter(max_retries=3))  # 访问https协议时，设置重传请求最多三次
    # 提取a标签的href和alt属性值
    category_data = [{'ppid': 12278, 'ppname': 'LDPE', 'navid': 521, 'name': 'PE'},
                     {'ppid': 12272, 'ppname': '共聚PP粒', 'navid': 525, 'name': 'PP粒'},
                     {'ppid': 12271, 'ppname': 'PP粉 ', 'navid': 526, 'name': ' PP粉'},
                     {'ppid': 12349, 'ppname': 'PVC粉', 'navid': 527, 'name': 'PVC'},
                     {'ppid': 12604, 'ppname': 'PS', 'navid': 531, 'name': 'PS'},
                     {'ppid': 12593, 'ppname': 'ABS', 'navid': 533, 'name': 'ABS'},
                     {'ppid': 12594, 'ppname': 'EPS普通料', 'navid': 529, 'name': 'EPS'},
                     {'ppid': 12276, 'ppname': 'EVA', 'navid': 534,
                      'name': 'EVA'}]  # 用于存储类别详情的列表，每个元素为一个字典，包含大类名称、大类URL、类别名称和类别URL

    # 使用正则表达式匹配类名
    for category in result:
        headers = {
            'Accept': '*/*',
            'Accept-Encoding': 'gzip, deflate, br',
            'Accept-Language': 'zh-CN,zh;q=0.9',
            'Content-Type': 'application/json',
            'Cookie': 'guid=236fbd2c-7794-b7d1-6943-950e0b0b033b; __bid_n=1877924aad211066474207; accessId=30dbced0-f5cb-11eb-893a-df95eeb4af27; Hm_lvt_44c27e8e603ca3b625b6b1e9c35d712d=1681364658,1681723276,1681870152; BAIDU_SSP_lcr=https://www.bing.com/; route=258ceb4bb660681c2cb2768af9756936; ASP.NET_SessionId=0snh542onvy54tecbm2g3z1b; Hm_lvt_78a951b1e2ee23efdc6af2ce70d6b9be=1681723475,1681796427,1681870273; href=https%3A%2F%2Fprices.sci99.com%2Fcn%2Fproduct.aspx%3Fppid%3D12278%26ppname%3DLDPE%26navid%3D521; sensorsdata2015jssdkcross=%7B%22distinct_id%22%3A%2218779249e7f58c-0ef11c4dd18171-26031851-2073600-18779249e80dbd%22%2C%22first_id%22%3A%22%22%2C%22props%22%3A%7B%22%24latest_traffic_source_type%22%3A%22%E8%87%AA%E7%84%B6%E6%90%9C%E7%B4%A2%E6%B5%81%E9%87%8F%22%2C%22%24latest_search_keyword%22%3A%22%E6%9C%AA%E5%8F%96%E5%88%B0%E5%80%BC%22%2C%22%24latest_referrer%22%3A%22https%3A%2F%2Fwww.bing.com%2F%22%7D%2C%22%24device_id%22%3A%2218779249e7f58c-0ef11c4dd18171-26031851-2073600-18779249e80dbd%22%7D; Hm_lpvt_44c27e8e603ca3b625b6b1e9c35d712d=1681902304; FPTOKEN=icFK4xODoKnx/04rfAoITwuUJBg68likg2sSI4JgBPdIv/S9jhgix39inzUaoGz5hAIiczPUUZQ2uJszR6v//D7+LcTqx40Z/CUIznbrR2yIMrHsmyatlhH0WgIhPLKznwqKkRvr+eQeiSkLZK189FdYUoll7FikJ6lja/uPw8ovlkQniYTMSypuNrT6+UTfSY3c/zNlDEAUfVlT3B6V+SjmLK99mEfkUVbWRRy1qhZiXReWpiXusi+ouWZe7ga+D3dvXgY0ykCQpbimCHhFwdtqP2VuF4NlgxYEuCrqlInCraqusNR1YVevG3KVdQyAOqCGSXMRebnNUhDRf80qOeIpN4W9diPx64vABA6OmttX0NMtsw4+0GqgYg77xHsn0eVDVaNA7bNo117fWDYJDA==|SCjqHrqxrSGlMmf5y8HpUsQThVuV6tMmVwf2Q0PT0AE=|10|2701e0700fa2e0fe5c4736e6b0d909ed; STATReferrerIndexId=2; Hm_lpvt_78a951b1e2ee23efdc6af2ce70d6b9be=1681952969; qimo_seosource_30dbced0-f5cb-11eb-893a-df95eeb4af27=%E5%85%B6%E4%BB%96%E7%BD%91%E7%AB%99; qimo_seokeywords_30dbced0-f5cb-11eb-893a-df95eeb4af27=%E6%9C%AA%E7%9F%A5; pageViewNum=300',
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/111.0.0.0 Safari/537.36'
        }
        body = {
            "region": "",
            "market": "",
            "factory": "",
            "model": "",
            "pname": "",
            "cycletype": "day",
            "pricecycle": "",
            "specialpricetype": "全部",
            "groupname": "",
            "ppname": category['name4'],
            "province": "",
            "pricetypeid": 34320,  # 34320 市场价格  34319企业报价
            "ppids": category['Ppid'],
            "navid": category['ID'],
            "sitetype": 1,
            "pageno": 1,
            "pagesize": 300,
            "purpose": ""
        }
        export_data = []
        res = requests.post('https://prices.sci99.com/api/zh-cn/product/datavalue', verify=False, timeout=100,
                            headers=headers, data=json.dumps(body))
        data_json = json.loads(res.text)
        headers = data_json['data']['headers']
        columns = ["", "", "", "", "", "", ""]
        time_columns = [header["Code"] for header in headers if
                        isinstance(header["Code"], str) and "/" in header["Code"] and datetime.strptime(header["Code"],
                                                                                                        "%Y/%m/%d")]
        new_list = [x for x in time_columns for _ in range(2)]
        columns.extend(new_list)

        items = data_json['data']['data']['Items']
        export_data = []
        for item in items:
            item_data = {}
            region = item.get("Region", "")
            province = item.get("Province", "")
            area = item.get("Area", "")
            if region == area:
                result = f"{region}-{province}" if province else region
            else:
                result = "-".join(filter(None, [region, province, area]))
                item_data['品名'] = category['name4']
                item_data['规格'] = item['Model']
                item_data['厂家/产地'] = result
                item_data['单位'] = item['Unit']
                item_data['备注'] = item['Remark']
                item_data['涨跌'] = item['Change']
                item_data['涨跌率'] = item['ChangeRate']

            for key in item.keys():
                if '/' in key:  # 时间键的特征是包含 /
                    price = item[key]
                    price_range = price.split('-')
                    if len(price_range) != 2:
                        continue
                    try:
                        min_price = float(price_range[0])
                        max_price = float(price_range[1])
                        item_data[key + '- 最低'] = min_price
                        item_data[key + '- 最高'] = max_price
                    except ValueError:
                        continue
            export_data.append(item_data)
        df = pd.DataFrame(export_data)
        # 计算时间范围
        end_date = datetime.now().strftime('%Y-%m-%d')  # 当前时间
        df.to_excel('D:\\工作记录\\dcost\\成本网站特点分析\\' + category['name2'] + '-' + category[
            'name4'] + '-' + end_date + '.xlsx', index=False)
        out_excel(category, data_json)


def out_excel(category, data):
    # 读取json数据

    # 创建工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    # 计算时间范围
    end_date = datetime.now().strftime('%Y-%m-%d')  # 当前时间
    # 写入表头
    headers = data['data']['headers']
    columns = ["", "", "", "", "", "", ""]
    time_columns = [header["Code"] for header in headers if
                    isinstance(header["Code"], str) and "/" in header["Code"] and datetime.strptime(header["Code"],
                                                                                                    "%Y/%m/%d")]
    new_list = [x for x in time_columns for _ in range(2)]
    columns.extend(new_list)
    two_columns = ['品名', '规格', '厂家/产地', '单位', '备注', '涨跌', '涨跌率']
    price_d = '最低价'
    for i, header in enumerate(columns):
        col = i + 1
        cell = ws.cell(row=1, column=col)
        cell.data_type = 'd'
        cell.number_format = 'yyyy/m/d'
        cell.alignment = Alignment(horizontal='left')  # 设置左对齐
        cell2 = ws.cell(row=2, column=col)
        if header != '':
            date_obj = datetime.strptime(header, '%Y/%m/%d')  # 将字符串转换为datetime对象
            cell.value = date_obj  # 将datetime对象写入单元格
        else:
            cell.value = None  # 将datetime对象写入单元格
        if header != '':
            cell2.value = price_d
            if price_d == '最高价':
                price_d = '最低价'
            else:
                price_d = '最高价'
        else:
            cell2.value = two_columns[i]

    # 写入数据
    items = data['data']['data']['Items']
    for i, item in enumerate(items):
        region = item.get("Region", "")
        province = item.get("Province", "")
        area = item.get("Area", "")
        if region == area:
            result = f"{region}-{province}" if province else region
        else:
            result = "-".join(filter(None, [region, province, area]))
        row = i + 3
        header_time = ''
        for j, header in enumerate(columns):
            col = j + 1
            if header != '':
                price = item.get(header)

                if price is not None:
                    price_range = price.split('-')
                    try:
                        min_price = float(price_range[0])
                        cell = ws.cell(row=row, column=col)
                        if header_time != header:
                            cell.value = min_price
                        else:
                            if len(price_range) == 2:
                                max_price = float(price_range[1])
                                cell.value = max_price
                            else:
                                cell.value=None
                        header_time = header

                    except ValueError:
                        continue

                else:

                    cell = ws.cell(row=row, column=col)
                    cell.value = None
                    continue
            else:
                cell = ws.cell(row=row, column=col)
                if j == 0:
                    cell.value = category['name4']
                if j == 1:
                    cell.value = item['Model']
                if j == 2:
                    cell.value = result
                if j == 3:
                    cell.value = item['Unit']
                if j == 4:
                    cell.value = item['Remark']
                if j == 5:
                    cell.value = item['Change']
                if j == 6:
                    cell.value = item['ChangeRate']
    # 保存文件
    wb.save(r'D:\工作记录\dcost\成本网站特点分析\塑料数据\excel\4444\\' + category['name2'] + '-' + category[
        'name4'] + '-' + end_date + '.xlsx')


if __name__ == '__main__':
    # print_hi('PyCharm')
    # print_csv()
    # print_csv1()
    # print_csv3()
    # print_csv4()
    #  print_save_sql()
    # print_excel_you_se()  # 有色金属
    # print_excel_xin_neng_yuan()  # 新能源
    # print_xin_nengyaun_guige()
    print_su_liao()  # 塑料数据收集

# See PyCharm help at https://www.jetbrains.com/help/pycharm/
